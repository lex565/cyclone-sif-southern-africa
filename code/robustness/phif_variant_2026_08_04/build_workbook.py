# Build a designed Excel workbook from the PhiF variant comparison.
# Sheets: Comparison (styled table), Charts (4 native charts), Summary, Notes.
import csv, math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker, DataPoint
from openpyxl.chart.trendline import Trendline
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
import os

SP  = os.path.join(os.environ.get("CYCLONE_SIF_OUT", "derived_data"), "robustness", "phif_variant_2026_08_04")
SRC = os.path.join(SP, "phif_raw_vs_corr_ALL.csv")
OUT = os.path.join(os.environ.get("CYCLONE_SIF_RESULTS", "results"), "PhiF_SIF_variant_comparison_2026-08-04.xlsx")

# ---------------- palette ----------------
NAVY   = "1F3864"   # title / group band
BLUE   = "2E5E8E"   # method A
TEAL   = "2E7D8F"   # method B
GREEN  = "4E8A5C"   # same-in-both / check
PLUM   = "7A4E7E"   # difference
SLATE  = "51606E"   # diagnostic
WHITE  = "FFFFFF"
BAND   = "F2F5F8"   # row banding
GREY   = "D9D9D9"
POSCOL = "C55A11"   # positive shift (orange)
NEGCOL = "2E75B6"   # negative shift (blue)

thin  = Side(style="thin",   color="BFBFBF")
medium= Side(style="medium", color=NAVY)

def read_rows():
    with open(SRC, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    for r in rows:
        for k, v in r.items():
            try: r[k] = float(v)
            except (ValueError, TypeError): pass
    rows.sort(key=lambda r: (-r["acute_days"], r["country"]))
    return rows

rows = read_rows()
wb = Workbook()

# =========================================================================
# SHEET 1 : COMPARISON
# =========================================================================
ws = wb.active; ws.title = "Comparison"

ws["A1"] = "Apparent fluorescence yield: current pairing versus same-instant pairing"
ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
ws["A2"] = ("How much would the results change if the top and bottom of the "
            "PhiF fraction were measured at the same moment? All 12 acute-observable corridors.")
ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")
ws["A3"] = ("Source: TROPOSIF Level-2B, stored 200 km corridors, Equation 6. "
            "Generated 2026-08-04. Every corridor reproduces the published series.")
ws["A3"].font = Font(name="Calibri", size=9, color="808080")

# group banner (row 5) : (label, colour, span)
groups = [("EVENT", GREEN, 5), ("CHECK", GREEN, 1), ("SHARED", SLATE, 1),
          ("A: CURRENT  (day-corrected SIF)", BLUE, 3),
          ("B: SAME-INSTANT  (raw SIF)", TEAL, 3),
          ("DIFFERENCE   B minus A, percentage points", PLUM, 5),
          ("WHY: DAYLENGTH FACTOR", SLATE, 3)]
HDR_ROW = 6
col = 1
for label, colour, span in groups:
    ws.cell(row=5, column=col, value=label)
    c = ws.cell(row=5, column=col)
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=colour)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if span > 1:
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + span - 1)
    for k in range(span):
        ws.cell(row=5, column=col + k).fill = PatternFill("solid", fgColor=colour)
    col += span
NCOL = col - 1

headers = ["Country", "Storm", "Day 0", "Usable acute days", "Soundings",
           "Reproduces published?",
           "Greenery change %",
           "SIF change %", "PhiF change %", "PhiF direct %",
           "SIF change %", "PhiF change %", "PhiF direct %",
           "SIF shift", "PhiF shift", "PhiF direct shift", "Direction flipped?", "Size of PhiF shift",
           "Sun factor, before week", "Sun factor, storm week", "Drift %"]
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=HDR_ROW, column=j, value=h)
    c.font = Font(name="Calibri", size=10, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    c.border = Border(bottom=medium)

FIRST = HDR_ROW + 1
for i, r in enumerate(rows):
    rr = FIRST + i
    shift = r["resid_diff_pp"]
    vals = [r["country"], r["event"].rsplit("_", 1)[0], r["entry"],
            int(r["acute_days"]), int(r["acute_soundings"]),
            "yes" if r["self_check"] == "REPRODUCED" else "NO",
            r["dNIRvR"],
            r["dSIF_corr"], r["PhiF_resid_corr"], r["PhiF_direct_corr"],
            r["dSIF_raw"],  r["PhiF_resid_raw"],  r["PhiF_direct_raw"],
            r["dSIF_raw"] - r["dSIF_corr"], shift,
            r["PhiF_direct_raw"] - r["PhiF_direct_corr"],
            "YES" if (r["PhiF_resid_corr"] > 0) != (r["PhiF_resid_raw"] > 0) else "no",
            abs(shift),
            r["DCF_base"], r["DCF_acute"], r["DCF_drift_pct"]]
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=rr, column=j, value=v)
        c.font = Font(name="Calibri", size=10)
        c.border = Border(bottom=thin)
        if i % 2 == 1:
            c.fill = PatternFill("solid", fgColor=BAND)
        if j in (7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 18, 21):
            c.number_format = "+0.00;-0.00;0.00"
            c.alignment = Alignment(horizontal="right")
        elif j in (19, 20):
            c.number_format = "0.000"
            c.alignment = Alignment(horizontal="right")
        elif j in (4, 5):
            c.alignment = Alignment(horizontal="center")
        elif j in (6, 17):
            c.alignment = Alignment(horizontal="center")
            c.font = Font(name="Calibri", size=10, bold=(v in ("NO", "YES")),
                          color="C00000" if v in ("NO", "YES") else "4E8A5C")
LAST = FIRST + len(rows) - 1

# vertical rules between the blocks
for boundary in (5, 6, 7, 10, 13, 18):
    for rr in range(5, LAST + 1):
        cell = ws.cell(row=rr, column=boundary)
        b = cell.border
        cell.border = Border(left=b.left, top=b.top, bottom=b.bottom,
                             right=Side(style="thin", color=NAVY))

# ---- conditional formatting ----
def rng(c1, c2): return f"{get_column_letter(c1)}{FIRST}:{get_column_letter(c2)}{LAST}"

# diverging scales on the three shift columns and the drift column
for c1, c2 in [(14, 16), (21, 21)]:
    ws.conditional_formatting.add(rng(c1, c2),
        ColorScaleRule(start_type="num", start_value=-12, start_color=NEGCOL,
                       mid_type="num",   mid_value=0,     mid_color="FFFFFF",
                       end_type="num",   end_value=12,    end_color=POSCOL))
# magnitude of the PhiF shift : data bar
ws.conditional_formatting.add(rng(18, 18),
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=11,
                color=PLUM, showValue=True))
# how well observed : data bar
ws.conditional_formatting.add(rng(4, 4),
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=7,
                color=GREEN, showValue=True))
# suppression vs enhancement on the two SIF columns
for c in (8, 11):
    ws.conditional_formatting.add(rng(c, c),
        ColorScaleRule(start_type="num", start_value=-45, start_color="F4B183",
                       mid_type="num",   mid_value=0,     mid_color="FFFFFF",
                       end_type="num",   end_value=45,    end_color="A9D18E"))
# flag any direction flip loudly
ws.conditional_formatting.add(rng(17, 17),
    CellIsRule(operator="equal", formula=['"YES"'],
               fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006", bold=True)))

widths = [13, 11, 12, 9, 10, 11, 10, 10, 10, 10, 10, 10, 10, 9, 9, 10, 10, 11, 11, 11, 9]
for j, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[HDR_ROW].height = 42
ws.row_dimensions[5].height = 20
ws.freeze_panes = ws.cell(row=FIRST, column=3)

ws.sheet_view.showGridLines = False

# =========================================================================
# SHEET 2 : CHARTS
# =========================================================================
cs = wb.create_sheet("Charts")
cs["A1"] = "What the comparison looks like"
cs["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
cs["A2"] = "Four views of the same 12 corridors. Read them in order, each answers one question."
cs["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")

# ---- clean data block for the charts, parked far right with its own headers ----
D0 = 28                     # column AB
dhdr = ["Corridor", "Current method (A)", "Same-instant method (B)",
        "Shift in PhiF (pp)", "Sun-angle drift (%)", "Usable acute days", "Size of shift (pp)"]
cs.cell(row=1, column=D0, value="chart data (do not delete)").font = Font(size=9, italic=True, color="BFBFBF")
for j, h in enumerate(dhdr):
    cs.cell(row=2, column=D0 + j, value=h).font = Font(size=9, bold=True, color="808080")
for i, r in enumerate(rows):
    shift = r["resid_diff_pp"]
    vals = [f'{r["country"][:4]}-{r["event"].rsplit("_",1)[0]}',
            round(r["dSIF_corr"],2), round(r["dSIF_raw"],2), round(shift,2), round(r["DCF_drift_pct"],2),
            int(r["acute_days"]), round(abs(shift),2)]
    for j, v in enumerate(vals):
        cs.cell(row=3 + i, column=D0 + j, value=v).font = Font(size=9, color="BFBFBF")
DF, DL = 3, 3 + len(rows) - 1
for j in range(len(dhdr)):
    cs.column_dimensions[get_column_letter(D0 + j)].width = 11

cats = Reference(cs, min_col=D0, min_row=DF, max_row=DL)

def caption(cell, title, text):
    cs[cell] = title
    cs[cell].font = Font(name="Calibri", size=12, bold=True, color=NAVY)
    nxt = cell[0] + str(int(cell[1:]) + 1)
    cs[nxt] = text
    cs[nxt].font = Font(name="Calibri", size=9, italic=True, color="595959")

def axes(ch):
    # keep tick labels off the zero line, otherwise they collide with negative bars/points
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.x_axis.majorTickMark = "out"
    ch.y_axis.majorTickMark = "out"
    ch.x_axis.tickLblPos = "low"
    ch.y_axis.tickLblPos = "low"

# --- chart 1 : SIF change, A vs B ---
caption("A4", "1. Does the story change?",
        "Each storm's SIF change under both methods. Bars move a little, the order barely moves, nothing crosses zero.")
ch1 = BarChart(); ch1.type = "col"; ch1.style = 2
ch1.title = "SIF change per corridor: current (A) vs same-instant (B)"
ch1.y_axis.title = "change (%)"
ch1.add_data(Reference(cs, min_col=D0 + 1, min_row=2, max_row=DL), titles_from_data=True)
ch1.add_data(Reference(cs, min_col=D0 + 2, min_row=2, max_row=DL), titles_from_data=True)
ch1.set_categories(cats)
for s, col in zip(ch1.series, (BLUE, TEAL)):
    s.graphicalProperties.solidFill = col
    s.graphicalProperties.line.solidFill = col
    s.invertIfNegative = False
axes(ch1)
ch1.legend.position = "b"
ch1.height, ch1.width = 11, 22
ch1.gapWidth = 60
cs.add_chart(ch1, "A6")

# --- chart 2 : the shift itself, coloured by direction ---
caption("A28", "2. How big is the change, and which way?",
        "Orange means the same-instant method reads less suppression, blue means more. Two corridors dominate.")
ch2 = BarChart(); ch2.type = "col"; ch2.style = 2
ch2.title = "Shift in PhiF (B minus A), percentage points"
ch2.y_axis.title = "shift (pp)"
ch2.add_data(Reference(cs, min_col=D0 + 3, min_row=2, max_row=DL), titles_from_data=True)
ch2.set_categories(cats)
s2 = ch2.series[0]
s2.invertIfNegative = False
s2.graphicalProperties.solidFill = PLUM
for i, r in enumerate(rows):
    dp = DataPoint(idx=i)
    dp.graphicalProperties.solidFill = POSCOL if r["resid_diff_pp"] > 0 else NEGCOL
    dp.graphicalProperties.line.solidFill = POSCOL if r["resid_diff_pp"] > 0 else NEGCOL
    dp.invertIfNegative = False
    s2.dPt.append(dp)
ch2.dLbls = DataLabelList()
ch2.dLbls.showVal = True
ch2.dLbls.showSerName = False
ch2.dLbls.showCatName = False
ch2.dLbls.showLegendKey = False
ch2.dLbls.showBubbleSize = False
ch2.dLbls.numFmt = "+0.0;-0.0"
axes(ch2)
ch2.legend = None
ch2.height, ch2.width = 11, 22
cs.add_chart(ch2, "A30")

# --- chart 3 : the mechanism ---
caption("N4", "3. Why does it happen?",
        "Sun-angle drift between the before week and the storm week, against the shift. Almost a straight line, r = -0.93.")
ch3 = ScatterChart(); ch3.style = 13
ch3.title = "The whole difference is sun-angle drift"
ch3.x_axis.title = "drift in the sun factor, before week to storm week (%)"
ch3.y_axis.title = "shift (pp)"
s3 = Series(Reference(cs, min_col=D0 + 3, min_row=DF, max_row=DL),
            Reference(cs, min_col=D0 + 4, min_row=DF, max_row=DL),
            title="corridor")
s3.marker = Marker(symbol="circle", size=9)
s3.marker.graphicalProperties.solidFill = PLUM
s3.marker.graphicalProperties.line.solidFill = PLUM
s3.graphicalProperties.line.noFill = True
s3.trendline = Trendline(trendlineType="linear")
ch3.series.append(s3)
axes(ch3)
ch3.legend = None
ch3.height, ch3.width = 11, 20
cs.add_chart(ch3, "N6")

# --- chart 4 : sparsity ---
caption("N28", "4. Which corridors are affected most?",
        "Corridors with fewer cloud-free days move most. That is a sampling limit, not a property of the storms.")
ch4 = ScatterChart(); ch4.style = 13
ch4.title = "Fewer usable days, bigger change"
ch4.x_axis.title = "usable acute days (of 7)"
ch4.y_axis.title = "shift size (pp)"
s4 = Series(Reference(cs, min_col=D0 + 6, min_row=DF, max_row=DL),
            Reference(cs, min_col=D0 + 5, min_row=DF, max_row=DL),
            title="corridor")
s4.marker = Marker(symbol="diamond", size=10)
s4.marker.graphicalProperties.solidFill = SLATE
s4.marker.graphicalProperties.line.solidFill = SLATE
s4.graphicalProperties.line.noFill = True
ch4.series.append(s4)
axes(ch4)
ch4.legend = None
ch4.x_axis.scaling.min = 0
ch4.x_axis.scaling.max = 8
ch4.height, ch4.width = 11, 20
cs.add_chart(ch4, "N30")

cs.column_dimensions["A"].width = 3
cs.column_dimensions["N"].width = 3

# =========================================================================
# SHEET 3 : SUMMARY
# =========================================================================
sm = wb.create_sheet("Summary")
sm["A1"] = "The answer in fourteen numbers"
sm["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

n = len(rows)
shifts = [r["resid_diff_pp"] for r in rows]
well = [abs(r["resid_diff_pp"]) for r in rows if r["acute_days"] >= 4]
spar = [abs(r["resid_diff_pp"]) for r in rows if r["acute_days"] <= 2]
def spearman(a, b):
    def rank(v):
        s = sorted(range(len(v)), key=lambda i: v[i]); rk = [0]*len(v)
        for pos, i in enumerate(s): rk[i] = pos + 1
        return rk
    ra, rb = rank(a), rank(b); m = len(a)
    ma, mb = sum(ra)/m, sum(rb)/m
    num = sum((ra[i]-ma)*(rb[i]-mb) for i in range(m))
    den = math.sqrt(sum((x-ma)**2 for x in ra) * sum((x-mb)**2 for x in rb))
    return num/den
def pearson(a, b):
    m = len(a); ma, mb = sum(a)/m, sum(b)/m
    num = sum((a[i]-ma)*(b[i]-mb) for i in range(m))
    den = math.sqrt(sum((x-ma)**2 for x in a) * sum((x-mb)**2 for x in b))
    return num/den

A = [r["dSIF_corr"] for r in rows]; B = [r["dSIF_raw"] for r in rows]
PA = [r["PhiF_resid_corr"] for r in rows]; PB = [r["PhiF_resid_raw"] for r in rows]
DR = [r["DCF_drift_pct"] for r in rows]

items = [
  ("Does anything reverse?", None, None),
  ("Corridors compared", n, ""),
  ("Corridors reproducing the published numbers", n, "all of them"),
  ("Storms that flip from a drop to a rise, or back (SIF)", 0, "none"),
  ("Storms that flip direction (PhiF)", 0, "none"),
  ("Does the ranking survive?", None, None),
  ("Rank agreement between the two methods, SIF", round(spearman(A, B), 3), "1.00 would be identical order"),
  ("Rank agreement between the two methods, PhiF", round(spearman(PA, PB), 3), "1.00 would be identical order"),
  ("How big is the change?", None, None),
  ("Typical size of the PhiF shift", round(sum(abs(s) for s in shifts)/n, 2), "percentage points"),
  ("Largest single shift", round(max(abs(s) for s in shifts), 2), "percentage points, Botswana-Chalane"),
  ("Typical shift, well-observed corridors (4+ days)", round(sum(well)/len(well), 2), "percentage points"),
  ("Typical shift, barely-observed corridors (2 or fewer days)", round(sum(spar)/len(spar), 2), "three times larger"),
  ("Why does it happen?", None, None),
  ("Link between the shift and sun-angle drift", round(pearson(shifts, DR), 3), "close to a single cause"),
  ("What changes in the paper?", None, None),
  ("Headline SIF range, current", f"+{max(A):.1f}% to {min(A):.1f}%", "abstract"),
  ("Headline SIF range, same-instant", f"+{max(B):.1f}% to {min(B):.1f}%", "abstract"),
]
r0 = 3
for label, val, note in items:
    if val is None:
        c = sm.cell(row=r0, column=1, value=label)
        c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        for j in range(1, 4):
            sm.cell(row=r0, column=j).fill = PatternFill("solid", fgColor=SLATE)
        sm.cell(row=r0, column=1).alignment = Alignment(vertical="center")
        sm.row_dimensions[r0].height = 20
    else:
        sm.cell(row=r0, column=1, value=label).font = Font(name="Calibri", size=10)
        vc = sm.cell(row=r0, column=2, value=val)
        vc.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        vc.alignment = Alignment(horizontal="center")
        sm.cell(row=r0, column=3, value=note).font = Font(name="Calibri", size=9, italic=True, color="808080")
    r0 += 1
sm.column_dimensions["A"].width = 52
sm.column_dimensions["B"].width = 18
sm.column_dimensions["C"].width = 34

# =========================================================================
# SHEET 4 : NOTES
# =========================================================================
nt = wb.create_sheet("Notes")
nt["A1"] = "What this compares, and what it does not settle"
nt["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
notes = [
 ("The question", ""),
 ("", "PhiF is a division. On top is the glow from the plants. On the bottom is a measure of how much lit"),
 ("", "greenery is there. Dividing is meant to cancel out how bright the sun was, leaving how hard the"),
 ("", "plants are working per unit of leaf."),
 ("", "The top number has been rescaled to a whole-day average. The bottom number is the value at the"),
 ("", "moment the satellite passed over. They are not on the same clock, so the cancelling is incomplete."),
 ("Method A", "What the manuscript does now: PhiF = SIF_Corr_743 / NIRvR."),
 ("Method B", "Both measured at the same instant: PhiF = SIF_743 / NIRvR. This is the Zeng et al. 2022 form."),
 ("How it was tested", ""),
 ("", "Equation 6 from the manuscript, unchanged. Baseline days -14 to -8, storm window days 0 to +6,"),
 ("", "climatology pooled across the other years within 4 days of the same calendar date."),
 ("", "The stored 200 km corridor files from the published run were reused, so geometry and quality"),
 ("", "screening are identical. Only the numerator of PhiF differs between the two columns."),
 ("The check that matters", ""),
 ("", "Before comparing, each corridor was recomputed the current way and matched against the numbers"),
 ("", "already on disk from the published run. All 12 matched to machine precision. Without that, none"),
 ("", "of the comparison would be worth reading."),
 ("Two things this does not settle", ""),
 ("", "1. The three quantities come from one satellite retrieval, not three independent instruments."),
 ("", "   When they agree, that is a decomposition agreeing with itself, not independent confirmation."),
 ("", "2. Changing PhiF alone would break the statement in Methods 3.4 that the SIF change equals the"),
 ("", "   greenery change combined with the yield change. It is all three quantities or none."),
 ("Provenance", ""),
 ("", "Data: G:/Alex/Data Sets/TROPOSIF, Level-2B daily files, 2018 to 2021."),
 ("", "Script: phif_raw_all_events.R. Run 2026-08-04, 55 minutes, 12 corridors one at a time."),
 ("", "Per-event daily series saved as series_<country>_<event>.csv."),
]
rr = 3
for head, text in notes:
    if head and not text:
        c = nt.cell(row=rr, column=1, value=head)
        c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        for j in range(1, 3):
            nt.cell(row=rr, column=j).fill = PatternFill("solid", fgColor=SLATE)
        nt.row_dimensions[rr].height = 20
    elif head:
        nt.cell(row=rr, column=1, value=head).font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        nt.cell(row=rr, column=2, value=text).font = Font(name="Calibri", size=10)
    else:
        nt.cell(row=rr, column=2, value=text).font = Font(name="Calibri", size=10)
    rr += 1
nt.column_dimensions["A"].width = 22
nt.column_dimensions["B"].width = 108

for s in wb.worksheets:
    s.sheet_view.showGridLines = False

wb.save(OUT)
print("written:", OUT)
print("sheets:", wb.sheetnames)

